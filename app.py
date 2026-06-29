import streamlit as st
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from fpdf import FPDF
from datetime import datetime, timezone, timedelta
import pandas as pd
import json
import qrcode
from io import BytesIO
from PIL import Image
from streamlit_drawable_canvas import st_canvas
import hashlib
import base64
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XlImage
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# Timezone WIB (UTC+7) untuk Indonesia Barat
WIB = timezone(timedelta(hours=7))

def now_wib():
    """Dapatkan waktu sekarang dalam timezone WIB (UTC+7)"""
    return datetime.now(WIB)

# Konfigurasi halaman
st.set_page_config(
    page_title="Sistem Absensi & Notulensi Rapat",
    page_icon="📋",
    layout="wide"
)

# Data hardcoded guru/staf (untuk admin)
DAFTAR_GURU = [
    {"nama": "Budi Santoso, S.Pd", "nip": "197501011998031001", "jabatan": "Guru Matematika"},
    {"nama": "Siti Nurhaliza, M.Pd", "nip": "198203052005012003", "jabatan": "Guru Bahasa Indonesia"},
    {"nama": "Ahmad Fauzi, S.Si", "nip": "198505152009031004", "jabatan": "Guru Biologi"},
    {"nama": "Dewi Lestari, S.Pd", "nip": "199002102012022001", "jabatan": "Guru Bahasa Inggris"},
    {"nama": "Eko Prasetyo, S.Kom", "nip": "199207202015031005", "jabatan": "Guru TIK"},
    {"nama": "Ratna Sari, S.Pd", "nip": "198808152010012008", "jabatan": "Guru PKn"},
    {"nama": "Muhammad Rizki, S.Pd", "nip": "199305102014031006", "jabatan": "Guru Olahraga"},
    {"nama": "Linda Wijaya, S.Pd", "nip": "199109252013022002", "jabatan": "Guru Seni Budaya"},
    {"nama": "Hendra Gunawan, M.Pd", "nip": "197212051997031002", "jabatan": "Wakil Kepala Sekolah"},
    {"nama": "Sri Mulyani, S.Pd", "nip": "198604182008012010", "jabatan": "Guru BK"}
]

# Fungsi koneksi Google Sheets
@st.cache_resource
def connect_to_gsheet():
    """Koneksi ke Google Sheets menggunakan credentials dari secrets"""
    try:
        credentials_dict = dict(st.secrets["gcp_service_account"])
        scope = [
            "https://spreadsheets.google.com/feeds",
            "https://www.googleapis.com/auth/drive"
        ]
        credentials = ServiceAccountCredentials.from_json_keyfile_dict(
            credentials_dict, scope
        )
        client = gspread.authorize(credentials)
        sheet = client.open_by_key(st.secrets["spreadsheet_key"])
        return sheet
    except Exception as e:
        error_msg = str(e)
        if "PERMISSION_DENIED" in error_msg or "403" in error_msg:
            st.error("❌ Google Sheets API atau Drive API belum di-enable! Buka Google Cloud Console → APIs & Services → Library → Enable 'Google Sheets API' dan 'Google Drive API'")
        elif "not found" in error_msg.lower() or "404" in error_msg:
            st.error("❌ Spreadsheet tidak ditemukan! Pastikan spreadsheet_key di secrets.toml sudah benar dan spreadsheet sudah di-share ke service account.")
        elif "invalid_grant" in error_msg.lower():
            st.error("❌ Credentials tidak valid! Pastikan private_key dan client_email sudah benar.")
        else:
            st.error(f"❌ Gagal koneksi ke Google Sheets: {error_msg if error_msg else 'Pastikan API sudah enable dan spreadsheet sudah di-share ke service account.'}")
        return None

def get_or_create_worksheet(sheet, worksheet_name, headers=None):
    """Ambil atau buat worksheet baru, otomatis tulis header jika belum ada"""
    created_new = False
    try:
        worksheet = sheet.worksheet(worksheet_name)
    except:
        worksheet = sheet.add_worksheet(title=worksheet_name, rows="1000", cols="20")
        created_new = True
    
    # Pastikan header ada di baris pertama
    if headers:
        try:
            first_row = worksheet.row_values(1)
            # Tulis header jika baris pertama kosong atau tidak cocok
            if not first_row or all(str(h).strip() == '' for h in first_row):
                worksheet.update('A1', [headers])
            else:
                # Cek apakah header yang ada cocok
                first_clean = [str(h).strip().lower() for h in first_row if str(h).strip()]
                expected_clean = [str(h).strip().lower() for h in headers]
                match = sum(1 for e in expected_clean if e in first_clean)
                if match == 0 and created_new:
                    worksheet.update('A1', [headers])
        except:
            worksheet.update('A1', [headers])
    
    return worksheet

def save_to_gsheet(worksheet, data):
    """Simpan data ke Google Sheets"""
    try:
        # Pastikan semua value jadi string untuk konsistensi
        data = [str(d) if d is not None else '' for d in data]
        worksheet.append_row(data, value_input_option='RAW')
        return True
    except Exception as e:
        st.error(f"Gagal menyimpan data: {str(e)}")
        return False

def update_row_in_gsheet(worksheet, row_index, data):
    """Update baris tertentu di Google Sheets (row_index 1-based, termasuk header)"""
    try:
        row_index = int(row_index)  # Pastikan Python int, bukan numpy int64
        data = [str(d) if d is not None else '' for d in data]
        for col_idx, value in enumerate(data, start=1):
            worksheet.update_cell(row_index, col_idx, value)
        return True
    except Exception as e:
        st.error(f"Gagal mengupdate data: {str(e)}")
        return False

def delete_row_in_gsheet(worksheet, row_index):
    """Hapus baris tertentu di Google Sheets (row_index 1-based, termasuk header)"""
    try:
        row_index = int(row_index)  # Pastikan Python int, bukan numpy int64
        worksheet.delete_rows(row_index)
        return True
    except Exception as e:
        st.error(f"Gagal menghapus data: {str(e)}")
        return False

def delete_rows_by_meeting_id(worksheet, meeting_id):
    """Hapus semua baris dengan meeting_id tertentu dari worksheet"""
    try:
        all_values = worksheet.get_all_values()
        if len(all_values) <= 1:
            return True
        # Cari dari bawah ke atas agar index tidak bergeser
        rows_to_delete = []
        for i in range(len(all_values) - 1, 0, -1):  # skip header (index 0)
            if str(all_values[i][0]).strip() == str(meeting_id).strip():
                rows_to_delete.append(i + 1)  # +1 karena gspread 1-based
        for row_idx in rows_to_delete:
            worksheet.delete_rows(row_idx)
        return True
    except Exception as e:
        st.error(f"Gagal menghapus data absensi: {str(e)}")
        return False

def generate_meeting_id():
    """Generate unique meeting ID"""
    timestamp = now_wib().strftime("%Y%m%d%H%M%S")
    return f"MTG{timestamp}"

def get_base_url():
    """Dapatkan base URL aplikasi secara otomatis"""
    # Cek apakah ada app_url di secrets
    app_url = st.secrets.get('app_url', '').strip().rstrip('/')
    if app_url and app_url != 'http://localhost:8501':
        return app_url
    
    # Auto-detect dari session (untuk Streamlit Cloud)
    try:
        from streamlit.web.server.websocket_headers import _get_websocket_headers
        headers = _get_websocket_headers()
        if headers:
            host = headers.get('Host', '')
            if host:
                scheme = 'https' if '.streamlit.app' in host else 'http'
                return f"{scheme}://{host}"
    except:
        pass
    
    return 'https://websiterapatonline.streamlit.app'

def generate_qr_code(meeting_id):
    """Generate QR Code untuk link absensi"""
    base_url = get_base_url()
    url = f"{base_url}?page=absensi&meeting_id={meeting_id}"
    
    qr = qrcode.QRCode(version=1, box_size=10, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    
    img = qr.make_image(fill_color="black", back_color="white")
    
    # Convert to bytes
    buf = BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    
    return buf, url

class PDFNotulensi(FPDF):
    """Class untuk generate PDF Notulensi Rapat"""
    
    def header(self):
        self.set_font('Arial', '', 12)
        self.cell(0, 6, 'PEMERINTAH KABUPATEN SIDOARJO', 0, 1, 'C')
        self.set_font('Arial', '', 12)
        self.cell(0, 6, 'DINAS PENDIDIKAN DAN KEBUDAYAAN', 0, 1, 'C')
        self.set_font('Arial', 'B', 16)
        self.cell(0, 8, 'SD NEGERI SIMOANGIN-ANGIN', 0, 1, 'C')
        self.set_font('Arial', '', 10)
        self.cell(0, 5, 'Jalan Simoangin-angin, Wonoayu, Sidoarjo, Jawa Timur 61261', 0, 1, 'C')
        self.cell(0, 5, 'Pos-el: sdnsimoangin@gmail.com', 0, 1, 'C')
        self.ln(2)
        self.set_line_width(0.8)
        self.line(10, self.get_y(), 200, self.get_y())
        self.set_line_width(0.3)
        self.line(10, self.get_y() + 1.5, 200, self.get_y() + 1.5)
        self.ln(6)
    
    def footer(self):
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Halaman {self.page_no()}', 0, 0, 'C')

def generate_pdf(data_rapat, peserta_list, notulensi):
    """Generate PDF Notulensi Rapat dengan daftar hadir lengkap"""
    
    pdf = PDFNotulensi()
    pdf.add_page()
    
    # Judul Dokumen
    pdf.set_font('Arial', 'B', 14)
    pdf.cell(0, 10, 'NOTULENSI RAPAT', 0, 1, 'C')
    pdf.ln(5)
    
    # Detail Rapat
    pdf.set_font('Arial', 'B', 11)
    pdf.cell(40, 7, 'DETAIL RAPAT', 0, 1)
    pdf.set_font('Arial', '', 10)
    
    details = [
        ('Meeting ID', data_rapat['meeting_id']),
        ('Judul Rapat', data_rapat['judul']),
        ('Tanggal', data_rapat['tanggal']),
        ('Waktu', data_rapat['waktu']),
        ('Lokasi', data_rapat['lokasi']),
        ('Pimpinan Rapat', data_rapat['pimpinan'])
    ]
    
    for label, value in details:
        pdf.cell(50, 6, f'{label}:', 0, 0)
        pdf.cell(0, 6, value, 0, 1)
    
    pdf.ln(5)
    
    # Daftar Hadir
    pdf.set_font('Arial', 'B', 11)
    pdf.cell(0, 7, f'DAFTAR HADIR ({len(peserta_list)} Peserta)', 0, 1)
    
    # Tabel daftar hadir
    pdf.set_font('Arial', 'B', 9)
    pdf.set_fill_color(200, 220, 255)
    pdf.cell(10, 7, 'No', 1, 0, 'C', True)
    pdf.cell(50, 7, 'Nama', 1, 0, 'C', True)
    pdf.cell(35, 7, 'NIP', 1, 0, 'C', True)
    pdf.cell(40, 7, 'Waktu Absen', 1, 0, 'C', True)
    pdf.cell(20, 7, 'Status', 1, 0, 'C', True)
    pdf.cell(35, 7, 'Tanda Tangan', 1, 1, 'C', True)
    
    pdf.set_font('Arial', '', 8)
    
    import tempfile, os
    temp_sig_files = []
    
    for idx, peserta in enumerate(peserta_list, 1):
        row_height = 15  # Tinggi baris untuk menampung TTD
        y_before = pdf.get_y()
        x_before = pdf.get_x()
        
        pdf.cell(10, row_height, str(idx), 1, 0, 'C')
        pdf.cell(50, row_height, peserta.get('Nama', ''), 1, 0)
        pdf.cell(35, row_height, peserta.get('NIP', ''), 1, 0)
        pdf.cell(40, row_height, peserta.get('Timestamp', ''), 1, 0)
        pdf.cell(20, row_height, 'Hadir', 1, 0, 'C')
        
        # Render TTD dari base64
        sig_data = peserta.get('Signature', '')
        sig_x = pdf.get_x()
        sig_y = pdf.get_y()
        
        if sig_data and len(sig_data) > 200:
            try:
                sig_bytes = base64.b64decode(sig_data)
                sig_img = Image.open(BytesIO(sig_bytes))
                # Simpan ke file temporary
                tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.png')
                sig_img.save(tmp.name, format='PNG')
                tmp.close()
                temp_sig_files.append(tmp.name)
                
                pdf.cell(35, row_height, '', 1, 1)  # Border kosong untuk kolom TTD
                # Gambar TTD di dalam sel
                pdf.image(tmp.name, x=sig_x + 2, y=sig_y + 1, w=30, h=row_height - 2)
            except:
                pdf.cell(35, row_height, '-', 1, 1, 'C')
        else:
            pdf.cell(35, row_height, '-', 1, 1, 'C')
    
    # Bersihkan file temporary
    for f in temp_sig_files:
        try:
            os.unlink(f)
        except:
            pass
    
    pdf.ln(5)
    
    # Isi Notulensi
    pdf.set_font('Arial', 'B', 11)
    pdf.cell(0, 7, 'ISI NOTULENSI', 0, 1)
    pdf.set_font('Arial', '', 10)
    pdf.multi_cell(0, 6, notulensi)
    
    pdf.ln(10)
    
    # Tanda tangan
    pdf.set_font('Arial', '', 10)
    pdf.cell(95, 6, f'Sidoarjo, {data_rapat["tanggal"]}', 0, 1)
    pdf.cell(95, 6, 'Notulis,', 0, 0)
    pdf.cell(95, 6, 'Mengetahui,', 0, 1)
    pdf.ln(15)
    
    pdf.set_font('Arial', 'U', 10)
    pdf.cell(95, 6, '(____________________)', 0, 0, 'C')
    pdf.cell(95, 6, data_rapat['pimpinan'], 0, 1, 'C')
    
    pdf.set_font('Arial', '', 9)
    pdf.cell(95, 5, '', 0, 0)
    pdf.cell(95, 5, 'Kepala Sekolah', 0, 1, 'C')
    
    # Output PDF
    timestamp = now_wib().strftime("%Y%m%d_%H%M%S")
    filename = f"Notulensi_Rapat_{timestamp}.pdf"
    pdf.output(filename)
    
    return filename

# ============= HALAMAN ADMIN =============
def admin_page():
    """Halaman Admin untuk membuat rapat dan generate link"""
    
    st.title("👨‍💼 Admin - Buat Rapat & Generate Link Absensi")
    st.markdown("---")
    
    # Sidebar info
    with st.sidebar:
        st.markdown("### SD NEGERI SIMOANGIN-ANGIN")
        st.markdown("Kab. Sidoarjo, Jawa Timur")
        st.info("**Mode: Admin**\nBuat rapat dan bagikan link absensi ke peserta.")
        
        st.markdown("---")
        st.markdown("**Status Koneksi:**")
        sheet = connect_to_gsheet()
        if sheet:
            st.success("✅ Terhubung ke Google Sheets")
        else:
            st.error("❌ Gagal terhubung")
    
    tab1, tab2, tab3, tab4 = st.tabs(["📝 Buat Rapat Baru", "📊 Lihat Daftar Hadir", "📄 Generate Notulensi", "✏️ Edit/Hapus Rapat"])
    
    # TAB 1: Buat Rapat
    with tab1:
        st.header("1️⃣ Data Rapat")
        col1, col2 = st.columns(2)
        
        with col1:
            judul_rapat = st.text_input(
                "Judul Rapat *",
                placeholder="Contoh: Rapat Koordinasi Semester Genap"
            )
            
            tanggal_rapat = st.date_input(
                "Tanggal Rapat *",
                value=now_wib()
            )
            
            lokasi_rapat = st.text_input(
                "Lokasi Rapat *",
                placeholder="Contoh: Ruang Guru"
            )
        
        with col2:
            waktu_rapat = st.time_input(
                "Waktu Rapat *",
                value=now_wib().time()
            )
            
            pimpinan_rapat = st.text_input(
                "Pimpinan Rapat *",
                placeholder="Contoh: Drs. Bambang Sutopo, M.Pd"
            )
        
        st.markdown("---")
        
        if st.button("🚀 Buat Rapat & Generate Link", type="primary", use_container_width=True):
            if not all([judul_rapat, lokasi_rapat, pimpinan_rapat]):
                st.error("❌ Semua field bertanda * wajib diisi!")
            else:
                with st.spinner("🔄 Sedang membuat rapat..."):
                    meeting_id = generate_meeting_id()
                    
                    # Simpan data rapat
                    if sheet:
                        rapat_headers = [
                            "Meeting ID", "Judul", "Tanggal", "Waktu", "Lokasi", 
                            "Pimpinan", "Timestamp Dibuat", "Status"
                        ]
                        worksheet_rapat = get_or_create_worksheet(sheet, "Data_Rapat", headers=rapat_headers)
                        
                        row_data = [
                            meeting_id,
                            judul_rapat,
                            tanggal_rapat.strftime("%d-%m-%Y"),
                            waktu_rapat.strftime("%H:%M"),
                            lokasi_rapat,
                            pimpinan_rapat,
                            now_wib().strftime("%Y-%m-%d %H:%M:%S"),
                            "Aktif"
                        ]
                        
                        if save_to_gsheet(worksheet_rapat, row_data):
                            st.success(f"✅ Rapat berhasil dibuat dengan ID: **{meeting_id}**")
                            
                            # Generate QR Code
                            qr_img, url = generate_qr_code(meeting_id)
                            
                            st.markdown("---")
                            st.success("### 🎉 Link Absensi Berhasil Dibuat!")
                            
                            col1, col2 = st.columns([1, 1])
                            
                            with col1:
                                st.markdown("#### 📱 QR Code")
                                st.image(qr_img, width=300)
                            
                            with col2:
                                st.markdown("#### 🔗 Link Absensi")
                                st.code(url, language="text")
                                st.info("**Cara penggunaan:**\n1. Bagikan link atau QR code ke peserta\n2. Peserta scan QR atau buka link\n3. Peserta isi nama, NIP, dan tanda tangan\n4. Data otomatis tersimpan ke Google Sheets")
                            
                            st.balloons()
    
    # TAB 2: Lihat Daftar Hadir
    with tab2:
        st.header("📊 Lihat Daftar Hadir Rapat")
        
        if sheet:
            try:
                absensi_headers = ["Meeting ID", "Nama", "NIP", "Timestamp", "Signature"]
                worksheet_absensi = get_or_create_worksheet(sheet, "Data_Absensi", headers=absensi_headers)
                df = read_sheet_as_dataframe(worksheet_absensi, expected_headers=absensi_headers)
                
                if not df.empty:
                    mid_col = find_column(df, 'Meeting ID')
                    if mid_col is None:
                        mid_col = df.columns[0]
                    df[mid_col] = df[mid_col].astype(str).str.strip()
                    
                    # Filter berdasarkan Meeting ID
                    meeting_ids = df[mid_col].unique()
                    selected_meeting = st.selectbox("Pilih Rapat:", meeting_ids)
                    
                    if selected_meeting:
                        df_filtered = df[df[mid_col] == selected_meeting]
                        
                        st.metric("Total Peserta Hadir", len(df_filtered))
                        
                        # Cari kolom
                        nama_col = find_column(df, 'Nama') or 'Nama'
                        nip_col = find_column(df, 'NIP') or 'NIP'
                        ts_col = find_column(df, 'Timestamp') or 'Timestamp'
                        sig_col = find_column(df, 'Signature') or 'Signature'
                        
                        show_cols = [c for c in [nama_col, nip_col, ts_col] if c in df_filtered.columns]
                        st.dataframe(
                            df_filtered[show_cols],
                            use_container_width=True
                        )
                        
                        # Tampilkan TTD peserta
                        if sig_col in df_filtered.columns:
                            st.markdown("#### ✍️ Tanda Tangan Peserta")
                            cols_ttd = st.columns(3)
                            for i, (_, row) in enumerate(df_filtered.iterrows()):
                                sig_data = row.get(sig_col, '')
                                with cols_ttd[i % 3]:
                                    st.markdown(f"**{row.get(nama_col, '')}**")
                                    if sig_data and len(sig_data) > 200:
                                        try:
                                            sig_bytes = base64.b64decode(sig_data)
                                            sig_img = Image.open(BytesIO(sig_bytes))
                                            st.image(sig_img, width=200)
                                        except:
                                            st.caption("TTD tidak dapat ditampilkan")
                                    else:
                                        st.caption("TTD tidak tersedia")
                        
                        st.markdown("---")
                        
                        # Download Excel dengan TTD
                        def generate_excel_daftar_hadir(df_data, meeting_id_str):
                            """Generate file Excel dengan kolom terpisah dan gambar TTD"""
                            wb = Workbook()
                            ws = wb.active
                            ws.title = "Daftar Hadir"
                            
                            # Style
                            header_font = Font(bold=True, color="FFFFFF", size=11)
                            header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
                            thin_border = Border(
                                left=Side(style='thin'), right=Side(style='thin'),
                                top=Side(style='thin'), bottom=Side(style='thin')
                            )
                            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                            
                            # Judul
                            ws.merge_cells('A1:E1')
                            ws['A1'] = f'DAFTAR HADIR RAPAT - {meeting_id_str}'
                            ws['A1'].font = Font(bold=True, size=14)
                            ws['A1'].alignment = Alignment(horizontal='center')
                            
                            # Header kolom di baris 3
                            headers = ['No', 'Nama', 'NIP', 'Waktu Absen', 'Tanda Tangan']
                            col_widths = [6, 30, 25, 22, 25]
                            
                            for col_idx, (header, width) in enumerate(zip(headers, col_widths), 1):
                                cell = ws.cell(row=3, column=col_idx, value=header)
                                cell.font = header_font
                                cell.fill = header_fill
                                cell.border = thin_border
                                cell.alignment = center_align
                                ws.column_dimensions[get_column_letter(col_idx)].width = width
                            
                            # Data
                            row_num = 4
                            for idx, (_, row) in enumerate(df_data.iterrows(), 1):
                                ws.row_dimensions[row_num].height = 60  # Tinggi untuk TTD
                                
                                # No
                                cell = ws.cell(row=row_num, column=1, value=idx)
                                cell.border = thin_border
                                cell.alignment = center_align
                                
                                # Nama
                                cell = ws.cell(row=row_num, column=2, value=str(row.get('Nama', '')))
                                cell.border = thin_border
                                cell.alignment = Alignment(vertical='center', wrap_text=True)
                                
                                # NIP
                                cell = ws.cell(row=row_num, column=3, value=str(row.get('NIP', '')))
                                cell.border = thin_border
                                cell.alignment = center_align
                                
                                # Timestamp
                                cell = ws.cell(row=row_num, column=4, value=str(row.get('Timestamp', '')))
                                cell.border = thin_border
                                cell.alignment = center_align
                                
                                # TTD
                                cell = ws.cell(row=row_num, column=5, value='')
                                cell.border = thin_border
                                
                                sig_data = row.get('Signature', '')
                                if sig_data and len(str(sig_data)) > 200:
                                    try:
                                        sig_bytes = base64.b64decode(str(sig_data))
                                        sig_stream = BytesIO(sig_bytes)
                                        sig_pil = Image.open(sig_stream)
                                        
                                        # Resize TTD
                                        sig_pil = sig_pil.resize((150, 50), Image.LANCZOS)
                                        
                                        img_buffer = BytesIO()
                                        sig_pil.save(img_buffer, format='PNG')
                                        img_buffer.seek(0)
                                        
                                        xl_img = XlImage(img_buffer)
                                        xl_img.width = 150
                                        xl_img.height = 50
                                        
                                        cell_ref = f'E{row_num}'
                                        ws.add_image(xl_img, cell_ref)
                                    except:
                                        ws.cell(row=row_num, column=5, value='(TTD tidak tersedia)')
                                else:
                                    ws.cell(row=row_num, column=5, value='(TTD tidak tersedia)')
                                
                                row_num += 1
                            
                            # Save ke BytesIO
                            output = BytesIO()
                            wb.save(output)
                            output.seek(0)
                            return output
                        
                        excel_data = generate_excel_daftar_hadir(df_filtered, selected_meeting)
                        st.download_button(
                            "📥 Download Daftar Hadir (Excel)",
                            excel_data,
                            f"daftar_hadir_{selected_meeting}.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                else:
                    st.info("Belum ada data absensi.")
            except Exception as e:
                st.error(f"Error: {str(e)}")
    
    # TAB 3: Generate Notulensi
    with tab3:
        st.header("📄 Generate Notulensi PDF")
        
        if sheet:
            try:
                rapat_headers = [
                    "Meeting ID", "Judul", "Tanggal", "Waktu", "Lokasi", 
                    "Pimpinan", "Timestamp Dibuat", "Status"
                ]
                worksheet_rapat = get_or_create_worksheet(sheet, "Data_Rapat", headers=rapat_headers)
                df_rapat = read_sheet_as_dataframe(worksheet_rapat, expected_headers=rapat_headers)
                
                if not df_rapat.empty:
                    mid_col_r = find_column(df_rapat, 'Meeting ID') or df_rapat.columns[0]
                    df_rapat[mid_col_r] = df_rapat[mid_col_r].astype(str).str.strip()
                    meeting_ids = df_rapat[mid_col_r].tolist()
                    
                    selected_meeting = st.selectbox("Pilih Rapat untuk Notulensi:", meeting_ids, key="notulensi")
                    
                    if selected_meeting:
                        rapat_data = df_rapat[df_rapat[mid_col_r] == selected_meeting].iloc[0]
                        
                        def rapat_val(col_name, default=''):
                            c = find_column(df_rapat, col_name)
                            if c and c in rapat_data.index:
                                v = rapat_data[c]
                                return str(v).strip() if v else default
                            return default
                        
                        st.info(f"**Judul:** {rapat_val('Judul')}\n\n**Tanggal:** {rapat_val('Tanggal')}\n\n**Pimpinan:** {rapat_val('Pimpinan')}")
                        
                        notulensi_text = st.text_area(
                            "Tulis Isi Notulensi *",
                            height=300,
                            placeholder="Tulis hasil pembahasan rapat..."
                        )
                        
                        if st.button("💾 Generate PDF Notulensi", type="primary"):
                            if notulensi_text:
                                # Ambil daftar hadir
                                absensi_headers = ["Meeting ID", "Nama", "NIP", "Timestamp", "Signature"]
                                worksheet_absensi = get_or_create_worksheet(sheet, "Data_Absensi", headers=absensi_headers)
                                df_absensi = read_sheet_as_dataframe(worksheet_absensi, expected_headers=absensi_headers)
                                
                                if not df_absensi.empty:
                                    abs_mid_col = find_column(df_absensi, 'Meeting ID') or df_absensi.columns[0]
                                    df_absensi[abs_mid_col] = df_absensi[abs_mid_col].astype(str).str.strip()
                                    peserta_df = df_absensi[df_absensi[abs_mid_col] == str(selected_meeting)]
                                    peserta_list = peserta_df.to_dict('records')
                                else:
                                    peserta_list = []
                                
                                data_rapat_dict = {
                                    'meeting_id': rapat_val('Meeting ID'),
                                    'judul': rapat_val('Judul'),
                                    'tanggal': rapat_val('Tanggal'),
                                    'waktu': rapat_val('Waktu'),
                                    'lokasi': rapat_val('Lokasi'),
                                    'pimpinan': rapat_val('Pimpinan')
                                }
                                
                                pdf_filename = generate_pdf(data_rapat_dict, peserta_list, notulensi_text)
                                
                                with open(pdf_filename, "rb") as pdf_file:
                                    st.download_button(
                                        "📥 Download PDF Notulensi",
                                        pdf_file.read(),
                                        pdf_filename,
                                        "application/pdf"
                                    )
                                
                                st.success("✅ PDF berhasil dibuat!")
                            else:
                                st.error("Notulensi harus diisi!")
                else:
                    st.info("Belum ada rapat yang dibuat.")
            except Exception as e:
                st.error(f"Error: {str(e)}")

    # TAB 4: Edit/Hapus Rapat
    with tab4:
        st.header("✏️ Kelola Rapat")
        
        if sheet:
            try:
                rapat_headers = [
                    "Meeting ID", "Judul", "Tanggal", "Waktu", "Lokasi",
                    "Pimpinan", "Timestamp Dibuat", "Status"
                ]
                worksheet_rapat = get_or_create_worksheet(sheet, "Data_Rapat", headers=rapat_headers)
                df_rapat_edit = read_sheet_as_dataframe(worksheet_rapat, expected_headers=rapat_headers)
                
                # Debug: tampilkan data mentah untuk diagnosis
                with st.expander("🔍 Debug: Data Mentah (klik untuk lihat)"):
                    raw_data = worksheet_rapat.get_all_values()
                    st.write(f"Total baris di sheet: {len(raw_data)}")
                    if raw_data:
                        st.write(f"Header di sheet: {raw_data[0]}")
                        st.write(f"Jumlah kolom header: {len(raw_data[0])}")
                        if len(raw_data) > 1:
                            st.write(f"Baris data pertama: {raw_data[1]}")
                            st.write(f"Jumlah kolom data: {len(raw_data[1])}")
                    st.write(f"DataFrame columns: {list(df_rapat_edit.columns)}")
                    st.write(f"DataFrame shape: {df_rapat_edit.shape}")
                    if not df_rapat_edit.empty:
                        st.dataframe(df_rapat_edit)
                
                if df_rapat_edit.empty:
                    st.info("Belum ada rapat yang dibuat.")
                else:
                    # Pastikan kolom Meeting ID ada
                    meeting_col = find_column(df_rapat_edit, "Meeting ID")
                    if meeting_col is None:
                        meeting_col = df_rapat_edit.columns[0]
                    
                    df_rapat_edit[meeting_col] = df_rapat_edit[meeting_col].astype(str).str.strip()
                    
                    # Tampilkan daftar rapat
                    st.subheader("📋 Daftar Rapat")
                    
                    display_cols = []
                    for col_name in ['Meeting ID', 'Judul', 'Tanggal', 'Waktu', 'Lokasi', 'Pimpinan', 'Status']:
                        found = find_column(df_rapat_edit, col_name)
                        if found:
                            display_cols.append(found)
                    
                    if display_cols:
                        st.dataframe(df_rapat_edit[display_cols], use_container_width=True)
                    
                    st.markdown("---")
                    
                    meeting_ids_edit = df_rapat_edit[meeting_col].tolist()
                    
                    # Helper untuk ambil nilai kolom aman
                    def get_col_val(row, col_name, default=''):
                        col = find_column(df_rapat_edit, col_name)
                        if col and col in row.index:
                            val = row[col]
                            return str(val).strip() if val else default
                        return default
                    
                    # Buat label yang informatif untuk selectbox
                    meeting_labels = []
                    for _, row in df_rapat_edit.iterrows():
                        mid = get_col_val(row, 'Meeting ID')
                        judul = get_col_val(row, 'Judul')
                        tgl = get_col_val(row, 'Tanggal')
                        label = f"{mid} - {judul} ({tgl})"
                        meeting_labels.append(label)
                    
                    selected_label = st.selectbox(
                        "Pilih Rapat untuk Edit/Hapus:",
                        meeting_labels,
                        key="edit_rapat_select"
                    )
                    
                    # Ambil meeting ID dari label
                    selected_idx = meeting_labels.index(selected_label)
                    selected_mid = meeting_ids_edit[selected_idx]
                    
                    # Cari data rapat yang dipilih
                    rapat_row = df_rapat_edit[df_rapat_edit[meeting_col] == selected_mid].iloc[0]
                    # Cari row index di Google Sheets (1-based + header row)
                    rapat_row_idx = df_rapat_edit[df_rapat_edit[meeting_col] == selected_mid].index[0] + 2
                    
                    # ---- SECTION EDIT ----
                    st.subheader("✏️ Edit Data Rapat")
                    
                    col_e1, col_e2 = st.columns(2)
                    
                    with col_e1:
                        edit_judul = st.text_input(
                            "Judul Rapat",
                            value=get_col_val(rapat_row, 'Judul'),
                            key="edit_judul"
                        )
                        try:
                            tgl_val = datetime.strptime(get_col_val(rapat_row, 'Tanggal'), '%d-%m-%Y')
                        except:
                            tgl_val = now_wib()
                        edit_tanggal = st.date_input(
                            "Tanggal Rapat",
                            value=tgl_val,
                            key="edit_tanggal"
                        )
                        edit_lokasi = st.text_input(
                            "Lokasi Rapat",
                            value=get_col_val(rapat_row, 'Lokasi'),
                            key="edit_lokasi"
                        )
                    
                    with col_e2:
                        waktu_str = get_col_val(rapat_row, 'Waktu', '00:00')
                        try:
                            waktu_val = datetime.strptime(waktu_str, '%H:%M').time()
                        except:
                            waktu_val = now_wib().time()
                        edit_waktu = st.time_input(
                            "Waktu Rapat",
                            value=waktu_val,
                            key="edit_waktu"
                        )
                        edit_pimpinan = st.text_input(
                            "Pimpinan Rapat",
                            value=get_col_val(rapat_row, 'Pimpinan'),
                            key="edit_pimpinan"
                        )
                        status_options = ["Aktif", "Selesai", "Dibatalkan"]
                        current_status = get_col_val(rapat_row, 'Status', 'Aktif')
                        status_idx = status_options.index(current_status) if current_status in status_options else 0
                        edit_status = st.selectbox(
                            "Status Rapat",
                            status_options,
                            index=status_idx,
                            key="edit_status"
                        )
                    
                    if st.button("💾 Simpan Perubahan", type="primary", key="btn_save_edit"):
                        if not all([edit_judul, edit_lokasi, edit_pimpinan]):
                            st.error("❌ Judul, Lokasi, dan Pimpinan wajib diisi!")
                        else:
                            with st.spinner("🔄 Menyimpan perubahan..."):
                                updated_row = [
                                    selected_mid,
                                    edit_judul,
                                    edit_tanggal.strftime("%d-%m-%Y"),
                                    edit_waktu.strftime("%H:%M"),
                                    edit_lokasi,
                                    edit_pimpinan,
                                    get_col_val(rapat_row, 'Timestamp Dibuat'),
                                    edit_status
                                ]
                                if update_row_in_gsheet(worksheet_rapat, rapat_row_idx, updated_row):
                                    st.success(f"✅ Rapat **{selected_mid}** berhasil diupdate!")
                                    st.rerun()
                    
                    st.markdown("---")
                    
                    # ---- SECTION HAPUS ----
                    st.subheader("🗑️ Hapus Rapat")
                    st.warning(f"⚠️ Menghapus rapat **{selected_mid}** akan menghapus data rapat beserta seluruh data absensi peserta. Tindakan ini tidak dapat dibatalkan!")
                    
                    konfirmasi_hapus = st.text_input(
                        f"Ketik **{selected_mid}** untuk konfirmasi penghapusan:",
                        key="konfirmasi_hapus",
                        placeholder=f"Ketik {selected_mid} di sini"
                    )
                    
                    if st.button("🗑️ Hapus Rapat Permanen", type="primary", key="btn_hapus_rapat"):
                        if konfirmasi_hapus.strip() != selected_mid.strip():
                            st.error("❌ Konfirmasi tidak cocok! Ketik Meeting ID dengan benar.")
                        else:
                            with st.spinner("🔄 Menghapus rapat dan data absensi..."):
                                # Hapus data absensi terkait
                                absensi_headers = ["Meeting ID", "Nama", "NIP", "Timestamp", "Signature"]
                                worksheet_absensi = get_or_create_worksheet(sheet, "Data_Absensi", headers=absensi_headers)
                                delete_rows_by_meeting_id(worksheet_absensi, selected_mid)
                                
                                # Hapus data rapat
                                if delete_row_in_gsheet(worksheet_rapat, rapat_row_idx):
                                    st.success(f"✅ Rapat **{selected_mid}** dan seluruh data absensinya berhasil dihapus!")
                                    st.rerun()
            except Exception as e:
                st.error(f"Error: {str(e)}")
        else:
            st.error("❌ Tidak dapat terhubung ke Google Sheets.")

# ============= HELPER UNTUK BACA SHEET ROBUST =============
def read_sheet_as_dataframe(worksheet, expected_headers=None):
    """Baca worksheet sebagai DataFrame dengan robust header handling.
    Menggunakan get_all_values() untuk menghindari masalah get_all_records().
    Jika expected_headers diberikan dan header di sheet tidak cocok, gunakan expected_headers.
    """
    all_values = worksheet.get_all_values()
    
    if not all_values or len(all_values) < 1:
        return pd.DataFrame()
    
    # Baris pertama = header
    raw_headers = [str(h).strip() for h in all_values[0]]
    
    # Cek apakah header di sheet cocok dengan expected_headers
    use_expected = False
    if expected_headers:
        # Cek: apakah header kosong atau tidak mengandung satupun expected header?
        non_empty_headers = [h for h in raw_headers if h != '']
        if not non_empty_headers:
            use_expected = True
        else:
            # Cek apakah setidaknya 1 expected header ada di raw headers
            expected_lower = [h.strip().lower() for h in expected_headers]
            raw_lower = [h.strip().lower() for h in raw_headers]
            match_count = sum(1 for e in expected_lower if e in raw_lower)
            if match_count == 0:
                use_expected = True
    
    if use_expected and expected_headers:
        headers = expected_headers
        # Jika header baris pertama tidak cocok, semua baris termasuk baris 1 mungkin = data
        # Tapi biasanya baris 1 tetap header, jadi kita skip baris 1
        data_rows = all_values[1:]
    else:
        headers = raw_headers
        data_rows = all_values[1:]
    
    # Jika data kosong
    if not data_rows:
        return pd.DataFrame(columns=headers)
    
    # Pastikan semua row punya jumlah kolom yang sama dengan header
    cleaned_rows = []
    for row in data_rows:
        # Skip baris yang sepenuhnya kosong
        if all(str(cell).strip() == '' for cell in row):
            continue
        # Pad atau trim row agar sesuai jumlah header
        if len(row) < len(headers):
            row = list(row) + [''] * (len(headers) - len(row))
        elif len(row) > len(headers):
            row = list(row)[:len(headers)]
        cleaned_rows.append(row)
    
    if not cleaned_rows:
        return pd.DataFrame(columns=headers)
    
    return pd.DataFrame(cleaned_rows, columns=headers)

def find_column(df, target_name):
    """Cari kolom di DataFrame secara case-insensitive dan strip whitespace."""
    target_lower = target_name.strip().lower()
    for col in df.columns:
        if col.strip().lower() == target_lower:
            return col
    # Fallback: cari partial match
    for col in df.columns:
        if target_lower.replace(' ', '') in col.strip().lower().replace(' ', ''):
            return col
    return None

# ============= HALAMAN FORM ABSENSI =============
def absensi_page():
    """Halaman Form Absensi untuk Peserta"""
    
    # Ambil meeting_id dari URL parameter
    query_params = st.query_params
    
    # Coba berbagai cara mendapatkan meeting_id
    meeting_id = None
    try:
        meeting_id = query_params.get("meeting_id", None)
    except Exception:
        pass
    
    if not meeting_id:
        try:
            params_dict = dict(query_params)
            meeting_id = params_dict.get("meeting_id", None)
            # Handle jika value berupa list (Streamlit versi lama)
            if isinstance(meeting_id, list) and len(meeting_id) > 0:
                meeting_id = meeting_id[0]
        except Exception:
            pass
    
    st.title("📝 Form Absensi Rapat")
    st.markdown("---")
    
    if not meeting_id:
        st.error("❌ Link tidak valid! Meeting ID tidak ditemukan.")
        st.info("Silakan gunakan link yang dibagikan oleh admin.")
        return
    
    # Pastikan meeting_id string bersih
    meeting_id = str(meeting_id).strip()
    
    # Ambil data rapat
    sheet = connect_to_gsheet()
    if not sheet:
        st.error("❌ Tidak dapat terhubung ke database.")
        return
    
    try:
        rapat_headers = [
            "Meeting ID", "Judul", "Tanggal", "Waktu", "Lokasi", 
            "Pimpinan", "Timestamp Dibuat", "Status"
        ]
        worksheet_rapat = get_or_create_worksheet(sheet, "Data_Rapat", headers=rapat_headers)
        
        # Gunakan read_sheet_as_dataframe untuk robust reading
        df_rapat = read_sheet_as_dataframe(worksheet_rapat, expected_headers=rapat_headers)
        
        if df_rapat.empty:
            st.error("❌ Belum ada data rapat. Minta admin untuk membuat rapat terlebih dahulu.")
            return
        
        # Cari kolom Meeting ID secara flexible
        meeting_col = find_column(df_rapat, "Meeting ID")
        if meeting_col is None:
            # Fallback: gunakan kolom pertama
            meeting_col = df_rapat.columns[0]
        
        # Pastikan tipe data meeting_id cocok (string, stripped)
        df_rapat[meeting_col] = df_rapat[meeting_col].astype(str).str.strip()
        rapat = df_rapat[df_rapat[meeting_col] == meeting_id]
        
        if rapat.empty:
            st.error(f"❌ Rapat dengan ID **{meeting_id}** tidak ditemukan!")
            return
        
        rapat_info = rapat.iloc[0]
        
        # Mapping kolom berdasarkan index sebagai fallback
        # Header: Meeting ID(0), Judul(1), Tanggal(2), Waktu(3), Lokasi(4), Pimpinan(5), Timestamp(6), Status(7)
        col_index_map = {
            "meeting id": 0, "judul": 1, "tanggal": 2, "waktu": 3,
            "lokasi": 4, "pimpinan": 5, "timestamp dibuat": 6, "status": 7
        }
        
        # Helper untuk akses kolom aman dengan fallback index
        def safe_get(col_name, default=""):
            # Coba 1: cari berdasarkan nama kolom
            col = find_column(df_rapat, col_name)
            if col is not None and col in rapat_info.index:
                val = rapat_info[col]
                if val is not None and str(val).strip() != '':
                    return str(val).strip()
            
            # Coba 2: akses berdasarkan index kolom
            idx = col_index_map.get(col_name.strip().lower())
            if idx is not None and idx < len(rapat_info):
                val = rapat_info.iloc[idx]
                if val is not None and str(val).strip() != '':
                    return str(val).strip()
            
            return default
        
        # Tampilkan info rapat
        judul = safe_get("Judul", "Rapat")
        tanggal = safe_get("Tanggal", "-")
        waktu = safe_get("Waktu", "-")
        lokasi = safe_get("Lokasi", "-")
        
        st.success(f"### 📋 {judul}")
        col1, col2, col3 = st.columns(3)
        with col1:
            st.info(f"**📅 Tanggal:** {tanggal}")
        with col2:
            st.info(f"**⏰ Waktu:** {waktu}")
        with col3:
            st.info(f"**📍 Lokasi:** {lokasi}")
        
        st.markdown("---")
        
        # Form Absensi
        st.header("✍️ Isi Data Absensi")
        
        nama = st.text_input(
            "Nama Lengkap *",
            placeholder="Contoh: Budi Santoso, S.Pd"
        )
        
        nip = st.text_input(
            "NIP *",
            placeholder="Contoh: 197501011998031001"
        )
        
        st.markdown("### ✍️ Tanda Tangan Digital")
        st.info("Silakan tanda tangan di kotak di bawah ini menggunakan mouse/touchscreen")
        
        # Counter untuk reset canvas (mengubah key agar canvas di-remount)
        if "canvas_key_counter" not in st.session_state:
            st.session_state.canvas_key_counter = 0
        
        # Canvas untuk tanda tangan
        canvas_result = st_canvas(
            stroke_width=3,
            stroke_color="#000000",
            background_color="#FFFFFF",
            height=200,
            width=600,
            drawing_mode="freedraw",
            key=f"signature_canvas_{st.session_state.canvas_key_counter}",
        )
        
        col1, col2 = st.columns([1, 3])
        
        with col1:
            if st.button("🔄 Hapus Tanda Tangan"):
                st.session_state.canvas_key_counter += 1
                st.rerun()
        
        st.markdown("---")
        
        if st.button("✅ Submit Absensi", type="primary", use_container_width=True):
            if not nama or not nip:
                st.error("❌ Nama dan NIP wajib diisi!")
            elif canvas_result.image_data is None or canvas_result.image_data.sum() == 0:
                st.error("❌ Tanda tangan belum dibuat!")
            else:
                # Cek duplikasi menggunakan read robust
                absensi_headers = ["Meeting ID", "Nama", "NIP", "Timestamp", "Signature"]
                worksheet_absensi = get_or_create_worksheet(sheet, "Data_Absensi", headers=absensi_headers)
                df_absensi = read_sheet_as_dataframe(worksheet_absensi, expected_headers=absensi_headers)
                
                if not df_absensi.empty:
                    abs_meeting_col = find_column(df_absensi, "Meeting ID")
                    abs_nip_col = find_column(df_absensi, "NIP")
                    
                    if abs_meeting_col and abs_nip_col:
                        df_absensi[abs_meeting_col] = df_absensi[abs_meeting_col].astype(str).str.strip()
                        df_absensi[abs_nip_col] = df_absensi[abs_nip_col].astype(str).str.strip()
                        sudah_absen = df_absensi[
                            (df_absensi[abs_meeting_col] == meeting_id) & 
                            (df_absensi[abs_nip_col] == str(nip).strip())
                        ]
                        
                        if not sudah_absen.empty:
                            st.warning("⚠️ Anda sudah melakukan absensi untuk rapat ini!")
                            return
                
                # Simpan tanda tangan sebagai base64
                img = Image.fromarray(canvas_result.image_data.astype('uint8'))
                buffered = BytesIO()
                img.save(buffered, format="PNG")
                signature_base64 = base64.b64encode(buffered.getvalue()).decode()
                
                row_data = [
                    meeting_id,
                    nama,
                    nip,
                    now_wib().strftime("%Y-%m-%d %H:%M:%S"),
                    signature_base64  # Simpan TTD lengkap sebagai base64
                ]
                
                if save_to_gsheet(worksheet_absensi, row_data):
                    st.success("✅ Absensi berhasil disimpan!")
                    st.balloons()
                    st.info("Terima kasih atas kehadiran Anda. Silakan tutup halaman ini.")
                else:
                    st.error("❌ Gagal menyimpan absensi. Silakan coba lagi.")
    
    except Exception as e:
        st.error(f"❌ Terjadi kesalahan: {str(e)}")
        st.info("💡 Pastikan admin sudah membuat rapat dan Google Sheets terhubung dengan benar.")

# ============= MAIN APP =============
def main():
    """Routing halaman"""
    
    query_params = st.query_params
    page = query_params.get("page", "admin")
    
    if page == "absensi":
        absensi_page()
    else:
        admin_page()
    
    # Footer
    st.markdown("---")
    st.markdown(
        "<div style='text-align: center; color: gray; font-size: 12px;'>"
        "© 2026 SD Negeri Simoangin-Angin | Sistem Absensi & Notulensi Rapat v2.0"
        "</div>",
        unsafe_allow_html=True
    )

if __name__ == "__main__":
    main()
